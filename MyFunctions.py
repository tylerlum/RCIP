# ---
# jupyter:
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: light
#       format_version: '1.5'
#       jupytext_version: 1.11.2
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

# +
from copy import deepcopy

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
blueFill = PatternFill(start_color='00FFFF',
                   end_color='00FFFF',
                   fill_type='solid')

# +
## Low-level functions ##
"""
Returns string of hh-mm-ss
"""
from time import gmtime, strftime
def getTime():
    return strftime("%H-%M-%S", gmtime())

"""
Splits input string by delimiter
Returns list of these strings 
"""
import re
def split(string):
    return re.split(', |; |\. |  ', string)

"""
Add column with to sheet to max column of sheet
Labels first row with column_header
"""
def add_column(sheet, column_header):
    max_col = sheet.max_column
    sheet.cell(row=1, column=max_col+1).value = column_header

"""
Returns True if statement both
1. Contains word
2. Has a letter immediately before or after
"""
def has_letter_before_or_after(statement, word):
    pattern1 = re.compile(".*\w+{}.*".format(word))
    pattern2 = re.compile(".*{}\w+.*".format(word))
    if pattern1.match(statement) or pattern2.match(statement):
        return True
    return False

"""
Returns True if statement contains any of words in keywords
Else return False
Also handles edge cases of short keywords
"""
def contains(statement, keywords):
    for word in keywords:
        ## Edge cases
        if word == "id":
            if not has_letter_before_or_after(statement, word) and word in statement:
                return True

        elif word == "ied":
            if not has_letter_before_or_after(statement, word) and word in statement:
                return True
            
        elif word == "tic":
            if not has_letter_before_or_after(statement, word) and word in statement:
                return True
                
        elif word == "si":
            if not has_letter_before_or_after(statement, word) and word in statement:
                return True
            
        elif word == "asd":
            if not has_letter_before_or_after(statement, word) and word in statement:
                return True
            
        elif word == "psychos":
            if word in statement:
                if not "psychosomatic" in statement: 
                    return True

        ## Regular word
        elif word in statement:
            return True

    return False

"""
Shortens diagnostic_type
admission => addx
discharge => dcdx
"""
def get_shortened_diagnostic_type(diagnostic_type):
    if diagnostic_type == "admission":
        return "addx"
    elif diagnostic_type == "discharge":
        return "dcdx"
    else:
        print("WARNING: BAD DIAGNOSTIC TYPE: {}".format(diagnostic_type))
        return ""

"""
Create a dictionary of column names
"""   
def get_column_names(sheet):
    column_names = {}
    i = 0
    for col in sheet.iter_cols(1, sheet.max_column):
        column_names[col[0].value] = i
        i += 1
    return column_names

"""
Create key list from dictionary
"""
def get_key_list(dictionary):
    key_list = []
    for i in dictionary.keys():
        key_list.append(i)
    return key_list

"""
Handle edge cases of illnesses
Return fixed illnesses list
"""
def manipulate_illness_list(illnesses, full_diagnosis):
    illnesses = deepcopy(illnesses)

    ## EOS/Personality
    if "Schizophrenia Spectrum and Other Psychotic Disorders" in illnesses and "Personality" in illnesses:
        illnesses.remove("Schizophrenia Spectrum and Other Psychotic Disorders")      
    
    return illnesses

"""
Input two lists
Output list containing all elements in long_list that are not in short_list
"""
def get_difference_list(short_list, long_list):
    short_list = deepcopy(short_list)
    long_list = deepcopy(long_list)

    for x in short_list:
        long_list.remove(x)

    return long_list

"""
Input list of strings
Split each string and add to list
Return list of all splits
"""
def combine_split_strings(strings):
    split_up_strings = []
    for string in strings:
        if type(string) is str:
            split_up_strings = split_up_strings + split(string.lower())
    return split_up_strings


# +
## Read-in Keywords from Sheets
"""
Read in sheet of illness keywords
Return dictionary mapping illnesses to list of keywords
"""
def read_in_illness_keywords(sheet):
    illness_keywords = {}
    sheet_columns = get_column_names(sheet)
    
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        illness = row_cells[sheet_columns['Illness']].value
        illness_keywords[illness] = []
        for col in range(sheet_columns['Illness']+1, sheet.max_column):
            if type(row_cells[col].value) is str:
                keyword = re.sub(r'\W+', '', row_cells[col].value)
                illness_keywords[illness].append(keyword.lower())
                
    return illness_keywords


"""
Read in sheet of reasons for referral keywords
Return dictionary mapping reasons for referral to list of keywrods
"""
def read_in_mrr_keywords(sheet):
    reasons_for_referral_keywords = {}
    sheet_columns = get_column_names(sheet)
    
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        reason_for_referral = row_cells[sheet_columns['Reasons']].value
        reasons_for_referral_keywords[reason_for_referral] = []
        for col in range(sheet_columns['Reasons']+1, sheet.max_column):
            if type(row_cells[col].value) is str:
                keyword = re.sub(r'\W+', '', row_cells[col].value)
                reasons_for_referral_keywords[reason_for_referral].append(keyword.lower())
    
    return reasons_for_referral_keywords

"""
Read in sheet of symptoms keywords
Return dictionary mapping symptoms to list of keywrods
"""
def read_in_symptom_keywords(sheet):
    symptom_keywords = {}
    sheet_columns = get_column_names(sheet)
    
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        symptom = row_cells[sheet_columns['Symptoms']].value
        symptom_keywords[symptom] = []
        for col in range(sheet_columns['Symptoms']+1, sheet.max_column):
            if type(row_cells[col].value) is str:
                keyword = re.sub(r'\W+', '', row_cells[col].value)
                symptom_keywords[symptom].append(keyword.lower())
                
    return symptom_keywords


# +
## Module 0: Add column headers
"""
Set column headers at end of columns
"""
def set_illness_headers(sheet, diagnostic_type, illnesses):
    ## Setup diagnostic type
    diagnostic_type_shortened = get_shortened_diagnostic_type(diagnostic_type)
    
    add_column(sheet, "main_{}".format(diagnostic_type_shortened))
    for illness in illnesses:
        add_column(sheet, "{}_{}".format(diagnostic_type_shortened, illness))

"""
Set column headers for referral reasons at end of columns
"""      
def set_mrr_headers(sheet, referral_reasons):
    for referral_reason in referral_reasons:
        add_column(sheet, "mrr_{}".format(referral_reason))

"""
Set column headers for symptoms at end of columns
"""      
def set_symptom_headers(sheet, symptoms):
    for symptom in symptoms:
        add_column(sheet, "{}".format(symptom))


# +
## Module 1: Set binaries to 0
"""
Set all columns with binary headers to 0
"""
def set_illnesses_to_zero(sheet, diagnostic_type, illnesses):
    ## Setup diagnostic type
    diagnostic_type_shortened = get_shortened_diagnostic_type(diagnostic_type)
    column_names = get_column_names(data_sheet)
    
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for illness in illnesses:
            row_cells[column_names["{}_{}".format(diagnostic_type_shortened, illness)]].value = 0

"""
Set all columns with binary headers to 0
"""            
def set_mrr_to_zero(sheet, referral_reasons):
    column_names = get_column_names(data_sheet)
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for referral_reason in referral_reasons:
            row_cells[column_names["mrr_{}".format(referral_reason)]].value = 0

"""
Set all columns with binary headers to 0
"""  
def set_symptoms_to_zero(sheet, symptoms):
    column_names = get_column_names(data_sheet)
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for symptom in symptoms:
            row_cells[column_names["{}".format(symptom)]].value = 0


# -

## Module 2: Fill in main diagnosis
"""
Find first illness in diagnosis, then label the main diagnosis
"""                           
def set_main_diagnosis(sheet, diagnostic_type, ignore_keywords, illness_keywords):
    ## Setup diagnostic type
    diagnostic_type_shortened = get_shortened_diagnostic_type(diagnostic_type)
    column_names = get_column_names(data_sheet)
    
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        
        ## Ensure valid cell data
        if type(row_cells[column_names['{}_diagnosis'.format(diagnostic_type)]].value) is str:

            ## Get diagnosis in lower-case
            full_diagnosis = row_cells[column_names['{}_diagnosis'.format(diagnostic_type)]].value.lower()
            diagnoses = split(full_diagnosis)
            illnesses = []

            first = True
            ## Loop through diagnoses until illness is found
            while len(diagnoses) > 0:
                main_diagnosis = diagnoses[0]

                ## Get main illness
                for illness in illness_keywords:
                    if not contains(main_diagnosis, ignore_keywords) and contains(main_diagnosis, illness_keywords[illness]):
                        illnesses.append(illness)
                    
                ## Update spreadsheet with illness
                if len(illnesses) == 1:
                    row_cells[column_names["main_{}".format(diagnostic_type_shortened)]].value = illnesses[0]
                    break
                    
                ## if more than 1, will show list in red
                elif len(illnesses) > 1:
                    
                    illnesses = manipulate_illness_list(illnesses, full_diagnosis)
                        
                    if len(illnesses) > 1:
                        row_cells[column_names["main_{}".format(diagnostic_type_shortened)]].value = str(illnesses)
                        row_cells[column_names["main_{}".format(diagnostic_type_shortened)]].fill = redFill
#                         print("***************")
#                         print(illnesses)
#                         print(main_diagnosis)
                    elif len(illnesses) == 1:
                        row_cells[column_names["main_{}".format(diagnostic_type_shortened)]].value = illnesses[0]
                    else:
                        print("EDGE CASE")

                    break
                if first:
                    first = False
#                     print("*******************")
#                     print(str(diagnoses))
                
                ## remove first element of diagnoses
                diagnoses = diagnoses[1:]
                
            ## If none found, fill with color
            if len(illnesses) == 0:
                row_cells[column_names["main_{}".format(diagnostic_type_shortened)]].fill = redFill


# +
## Module 3: Read diagnoses, One hot encode illnesses
"""
For each illness in diagnosis, label the corresponding column with a 1
"""
def ohe_illnesses(sheet, diagnostic_type, ignore_keywords, illness_keywords):
    ## Setup diagnostic type
    diagnostic_type_shortened = get_shortened_diagnostic_type(diagnostic_type)
    column_names = get_column_names(data_sheet)
    
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        
        ## Ensure valid cell data
        if type(row_cells[column_names['{}_diagnosis'.format(diagnostic_type)]].value) is str:
            
            ## Get diagnosis in lower-case
            full_diagnosis = row_cells[column_names['{}_diagnosis'.format(diagnostic_type)]].value.lower()
            diagnoses = split(full_diagnosis)
            illnesses = []
            
            ## Add illnesses for each statement in diagnosis
            for statement in diagnoses:
                    statement_illnesses = []
                    ## Add all illnesses associated with each statement
                    for illness in illness_keywords:
                        if not contains(statement, ignore_keywords) and contains(statement, illness_keywords[illness]):
                            statement_illnesses.append(illness)
                            row_cells[column_names['{}_{}'.format(diagnostic_type_shortened, illness)]].value = 1
                    illnesses = illnesses + statement_illnesses

#                    if len(statement_illnesses) > 1:
#                         print("*****************")
#                         print(statement_illnesses)
#                         print(statement)
                    ## Exceptions (REMOVES SUICIDE IF ONLY 2, should it?)
                    manipulated_illnesses = manipulate_illness_list(illnesses, full_diagnosis)
                    difference_list = get_difference_list(manipulated_illnesses, illnesses)
                    for difference in difference_list:
                        row_cells[column_names['{}_{}'.format(diagnostic_type_shortened, difference)]].value = 0


# -

## Module 4: Read ref_reason + chief_complaint, One hot encode main reason for referrals
"""
For each illness in diagnosis, label the corresponding column with a 1
"""
def ohe_mrr(sheet, ignore_keywords, reasons_for_referral_keywords):
    column_names = get_column_names(data_sheet)
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        
        ## Ensure valid cell data
        if type(row_cells[column_names['ref_reason']].value) is str:
            
            ## Get reason for referral in lower-case
            full_ref_reason = row_cells[column_names['ref_reason']].value.lower()
            ref_reasons = split(full_ref_reason)
            reasons = []
            
            ## Add reasons for each statement in ref_reasons
            for statement in ref_reasons:
                    statement_reasons = []
                    ## Add all reasons associated with each statement
                    for ref_reason in reasons_for_referral_keywords:
                        if not contains(statement, ignore_keywords) and contains(statement, reasons_for_referral_keywords[ref_reason]):
                            statement_reasons.append(ref_reason)
                            row_cells[column_names['mrr_{}'.format(ref_reason)]].value = 1
                    reasons = reasons + statement_reasons
    
    # Bad practice copy paste
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        
        ## Ensure valid cell data
        if type(row_cells[column_names['chief_complaint']].value) is str:
            
            ## Get reason for referral in lower-case
            full_ref_reason = row_cells[column_names['chief_complaint']].value.lower()
            ref_reasons = split(full_ref_reason)
            reasons = []
            
            ## Add reasons for each statement in ref_reasons
            for statement in ref_reasons:
                    statement_reasons = []
                    ## Add all reasons associated with each statement
                    for ref_reason in reasons_for_referral_keywords:
                        if not contains(statement, ignore_keywords) and contains(statement, reasons_for_referral_keywords[ref_reason]):
                            statement_reasons.append(ref_reason)
                            row_cells[column_names['mrr_{}'.format(ref_reason)]].value = 1
                    reasons = reasons + statement_reasons
    
    # Fill in other column
    reasons_for_referral = get_key_list(reasons_for_referral_keywords)
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        
        isOther = True
        for reason_for_referral in reasons_for_referral:
            if row_cells[column_names['mrr_{}'.format(reason_for_referral)]].value == 1:
                isOther = False
                break
        
        if isOther:
            row_cells[column_names['mrr_Other']].value = 1
#             print("Referral reason: {}".format(row_cells[column_names['ref_reason']].value))
#             print("Chief complaint: {}".format(row_cells[column_names['chief_complaint']].value))
#             print("**************************")


## Module 5: Read all four info columns, One hot encode symptoms
"""
For each symptom in diagnosis, label the corresponding column with a 1
"""
def ohe_symptoms(sheet, ignore_keywords, symptom_keywords):
    column_names = get_column_names(data_sheet)
    
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        
        ## Get all diagnosis info
        ref_reason = row_cells[column_names['ref_reason']].value
        chief_complaint = row_cells[column_names['chief_complaint']].value
        admission_diagnosis = row_cells[column_names['admission_diagnosis']].value
        discharge_diagnosis = row_cells[column_names['discharge_diagnosis']].value
        
        all_info = []
        all_info.extend((ref_reason, chief_complaint, admission_diagnosis, discharge_diagnosis))
        
        split_all_info = combine_split_strings(all_info)
        
        all_reasons = []
        for info in split_all_info:
            info_reasons = []
            ## add all symptoms associated with each statement
            for symptom in symptom_keywords:
                if not contains(info, ignore_keywords) and contains(info, symptom_keywords[symptom]):
                    info_reasons.append(symptom)
                    row_cells[column_names['{}'.format(symptom)]].value = 1
            all_reasons = all_reasons + info_reasons
